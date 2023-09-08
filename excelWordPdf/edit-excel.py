import openpyxl
import os

wb = openpyxl.Workbook() # create a new workbook
sheet = wb.get_sheet_by_name('Sheet') # get the sheet

sheet['A1'] = 42 # assign a value to a cell
sheet['A2'] = 'Hello' # assign a value to a cell

sheet2 = wb.create_sheet() # create a new sheet

sheet2.title = 'My New Sheet Name' # change the sheet name
wb.create_sheet(index=0, title='My Other Sheet') # create a new sheet at index 0 or the first sheet
print(wb.get_sheet_names()) # get the sheet names
# can import os and use os.chdir() to change the directory to where you want to save this file
cwd = os.getcwd() # get the current working directory
os.chdir('/home/<username>/<directory>') # change your username to your username
print(cwd) # print the current working directory
wb.save('example2.xlsx') # save the workbook ! update this as you make changes so you dont modify original file

