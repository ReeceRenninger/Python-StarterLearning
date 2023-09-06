import openpyxl
import os

os.chdir('home/<username>/example.xlsx')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

# grab the value in cell A1
cell = sheet['A1']
str(cell.value)

newCell = sheet['B1'].value
print(newCell)

# navigate to specific row or cell rather than calling the cell by name
sheet.cell(row=1, column=2)

# iterate through the rows using i as the row number
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
